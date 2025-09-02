import csv
import os
import psycopg2
import io
import time
import threading
from flask import Flask, jsonify, request, Response
from flask_cors import CORS
from dotenv import load_dotenv
import pandas as pd
import gspread
from google.oauth2.service_account import Credentials
from concurrent.futures import ThreadPoolExecutor, as_completed
from psycopg2 import pool
import openpyxl

# Load environment variables from .env file
load_dotenv()

# --- Database Connection Pool ---
db_pool = None

def init_db_pool():
    """Initializes the database connection pool."""
    global db_pool
    if db_pool:
        db_pool.closeall()
    try:
        db_pool = pool.ThreadedConnectionPool(
            minconn=1,
            maxconn=10,
            host=os.getenv('DB_HOST'),
            port=os.getenv('DB_PORT'),
            dbname=os.getenv('DB_NAME'),
            user=os.getenv('DB_USER'),
            password=os.getenv('DB_PASSWORD')
        )
        print("Database connection pool initialized successfully.")
        return True
    except psycopg2.Error as e:
        print(f"Error initializing database pool: {e}")
        db_pool = None
        return False

def get_db_connection():
    """Gets a connection from the pool."""
    if not db_pool:
        if not init_db_pool():
            raise ConnectionError("Database connection pool is not available.")
    return db_pool.getconn()

def return_db_connection(conn):
    """Returns a connection to the pool."""
    if db_pool:
        db_pool.putconn(conn)

# --- Flask App Initialization ---
app = Flask(__name__)
CORS(app)

# --- Error Handling ---
@app.errorhandler(Exception)
def handle_exception(e):
    """Generic error handler."""
    app.logger.error(f"An error occurred: {e}", exc_info=True)
    return jsonify(error=str(e)), 500

# --- API Endpoints ---

def _test_single_db_connection(host, port, dbname, user, password):
    """Attempts to establish a single database connection with provided parameters."""
    conn = None
    try:
        conn = psycopg2.connect(
            host=host, port=port, dbname=dbname, user=user, password=password, connect_timeout=5
        )
        return True, "Database connection successful!"
    except psycopg2.Error as e:
        return False, f"Database connection failed: {e}"
    finally:
        if conn:
            conn.close()

@app.route('/api/db/test', methods=['POST'])
def test_db_connection():
    """Tests the database connection using parameters from the request body."""
    config = request.json
    db_host = config.get('dbHost')
    db_port = config.get('dbPort')
    db_name = config.get('dbName')
    db_user = config.get('dbUser')
    db_password = config.get('dbPassword')

    if not all([db_host, db_port, db_name, db_user, db_password]):
        return jsonify(status='error', message='All database configuration fields must be provided.'), 400

    success, message = _test_single_db_connection(db_host, db_port, db_name, db_user, db_password)
    if success:
        return jsonify(status='success', message=message)
    else:
        return jsonify(status='error', message=message), 500

@app.route('/api/db/schema', methods=['POST'])
def create_schema_endpoint():
    """Endpoint to create the database schema."""
    conn = None
    log = []
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            log.append("Dropping existing schema objects if they exist...")
            cursor.execute("DROP TABLE IF EXISTS rings;")
            cursor.execute("DROP FUNCTION IF EXISTS update_rings_tsvector_trigger CASCADE;")

            log.append("Creating the 'rings' table and base indexes...")
            create_table_sql = """
            CREATE TABLE rings (
                id SERIAL PRIMARY KEY, date DATE, mo_number VARCHAR(50), vendor VARCHAR(50),
                serial_number VARCHAR(100) UNIQUE, ring_size VARCHAR(100), sku VARCHAR(50),
                vqc_status VARCHAR(100), vqc_reason TEXT, ft_status VARCHAR(100), ft_reason TEXT,
                reason_tsvector TSVECTOR, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );
            CREATE INDEX idx_serial_number ON rings(serial_number);
            CREATE INDEX idx_vendor ON rings(vendor);
            CREATE INDEX idx_date_desc ON rings(date DESC);
            """
            cursor.execute(create_table_sql)

            log.append("Adding optimized composite and full-text search indexes...")
            optimized_indexes_sql = """
            CREATE INDEX idx_rings_composite ON rings(vendor, vqc_status, ft_status);
            CREATE INDEX idx_rings_text_search ON rings USING GIN(reason_tsvector);
            """
            cursor.execute(optimized_indexes_sql)

            log.append("Creating trigger function for automatic full-text search indexing...")
            tsvector_trigger_sql = """
            CREATE OR REPLACE FUNCTION update_rings_tsvector_trigger() RETURNS trigger AS $$
            BEGIN
                NEW.reason_tsvector :=
                    to_tsvector('english', COALESCE(NEW.vqc_reason, '') || ' ' || COALESCE(NEW.ft_reason, ''));
                RETURN NEW;
            END;
            $$ LANGUAGE plpgsql;

            CREATE TRIGGER tsvectorupdate BEFORE INSERT OR UPDATE
            ON rings FOR EACH ROW EXECUTE PROCEDURE update_rings_tsvector_trigger();
            """
            cursor.execute(tsvector_trigger_sql)

        conn.commit()
        log.append("Database schema, optimized indexes, and triggers created successfully.")
        return jsonify(status="success", logs=log)
    except psycopg2.Error as db_err:
        if conn: conn.rollback()
        app.logger.error(f"Database error during schema creation: {db_err}")
        return jsonify(status="error", message=f"Database error during schema creation: {db_err}"), 500
    finally:
        if conn: return_db_connection(conn)


@app.route('/api/db/clear', methods=['DELETE'])
def clear_database_endpoint():
    """Endpoint to clear the 'rings' table."""
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute("TRUNCATE TABLE rings RESTART IDENTITY")
        conn.commit()
        return jsonify(status="success", message="Database 'rings' table has been cleared.")
    except (psycopg2.Error, Exception) as e:
        if conn: conn.rollback()
        app.logger.error(f"Database clearing failed: {e}")
        return jsonify(status="error", message=f"Database clearing failed: {e}"), 500
    finally:
        if conn: return_db_connection(conn)


@app.route('/api/data', methods=['GET'])
def get_data():
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute('SELECT * FROM rings;') # Corrected table name
    
    # Fetch column names from cursor description
    colnames = [desc[0] for desc in cur.description]
    
    # Fetch all rows and convert to list of dictionaries
    data = [dict(zip(colnames, row)) for row in cur.fetchall()]
    
    cur.close()
    conn.close()
    return jsonify(data)


# --- Helper Functions for Data Processing (from data_handler.py) ---

def _load_sheet_data(sheet_type, config, gc, log_callback):
    try:
        if sheet_type == 'step7':
            sheet = gc.open_by_url(config['vendorDataUrl'])
            ws = sheet.worksheet('Working') # Assuming worksheet name
            log_callback(f"Loading {sheet_type} data...")
            all_values = ws.get_all_values()
            if not all_values: return 'step7', []
            headers = [str(h).strip() if h else f"Empty_Col_{i}" for i, h in enumerate(all_values[0])]
            data = [dict(zip(headers, row)) for row in all_values[1:]]
            log_callback(f"Loaded {len(data)} records from {sheet_type}")
            return 'step7', data
        
        elif sheet_type == 'vqc':
            vqc_data = {}
            vqc_sheet = gc.open_by_url(config['vqcDataUrl'])
            log_callback("Loading VQC data...")
            # Assuming worksheet names are the vendor names
            for vendor in ['IHC', '3DE TECH', 'MAKENICA']:
                try:
                    ws = vqc_sheet.worksheet(vendor)
                    all_values = ws.get_all_values()
                    if all_values:
                        headers = [str(h).strip() if h else f"Empty_Col_{i}" for i, h in enumerate(all_values[0])]
                        vqc_data[vendor] = [dict(zip(headers, row)) for row in all_values[1:]]
                        log_callback(f"Loaded {len(vqc_data[vendor])} VQC records for {vendor}")
                except Exception as e:
                    log_callback(f"Warning: Could not load VQC sheet for '{vendor}': {e}")
            return 'vqc', vqc_data
        
        elif sheet_type == 'ft':
            sheet = gc.open_by_url(config['ftDataUrl'])
            ws = sheet.worksheet('Working') # Assuming worksheet name
            log_callback(f"Loading {sheet_type} data...")
            all_values = ws.get_all_values()
            if not all_values: return 'ft', []
            headers = [str(h).strip() if h else f"Empty_Col_{i}" for i, h in enumerate(all_values[0])]
            data = [dict(zip(headers, row)) for row in all_values[1:]]
            log_callback(f"Loaded {len(data)} FT records")
            return 'ft', data

    except Exception as e:
        log_callback(f"ERROR loading {sheet_type} data: {e}")
        if sheet_type == 'vqc': return 'vqc', {}
        return sheet_type, []

def _merge_ring_data_fast(step7_data, vqc_data, ft_data, log_callback):
    if not step7_data: return []

    def find_column(df, patterns):
        if not isinstance(patterns, list): patterns = [patterns]
        for pattern in patterns:
            for col in df.columns:
                if pattern.lower() == str(col).lower().strip():
                    return col
        return None

    log_callback("Reshaping main vendor data...")
    df_step7 = pd.DataFrame(step7_data)
    vendor_mappings = {
        '3DE TECH': {'serial': 'UID', 'mo': '3DE MO', 'sku': 'SKU', 'size': 'SIZE'},
        'IHC': {'serial': 'IHC', 'mo': 'IHC MO', 'sku': 'IHC SKU', 'size': 'IHC SIZE'},
        'MAKENICA': {'serial': 'MAKENICA', 'mo': 'MK MO', 'sku': 'MAKENICA SKU', 'size': 'MAKENICA SIZE'}
    }
    all_vendor_dfs = []
    date_col = find_column(df_step7, ['logged_timestamp', 'timestamp', 'date'])
    for vendor, patterns in vendor_mappings.items():
        serial_col = find_column(df_step7, patterns['serial'])
        if not serial_col: continue
        mo_col, sku_col, size_col = find_column(df_step7, patterns['mo']), find_column(df_step7, patterns['sku']), find_column(df_step7, patterns['size'])
        cols_to_keep = {date_col: 'date', serial_col: 'serial_number', mo_col: 'mo_number', sku_col: 'sku', size_col: 'ring_size'}
        cols_to_keep = {k: v for k, v in cols_to_keep.items() if k is not None and k in df_step7.columns}
        vendor_df = df_step7[list(cols_to_keep.keys())].copy()
        vendor_df.rename(columns=cols_to_keep, inplace=True)
        vendor_df['vendor'] = vendor
        if 'serial_number' in vendor_df.columns:
            vendor_df.dropna(subset=['serial_number'], inplace=True)
            vendor_df['serial_number'] = vendor_df['serial_number'].astype(str).str.strip()
            all_vendor_dfs.append(vendor_df[vendor_df['serial_number'] != ''])
        else:
            log_callback(f"WARNING: 'serial_number' column not found for vendor {vendor}. Skipping this vendor's data.")
    
    if not all_vendor_dfs: raise ValueError("Could not process any vendor data from Step 7.")
    df_main = pd.concat(all_vendor_dfs, ignore_index=True)
    log_callback(f"Reshaped into {len(df_main)} total records.")

    log_callback("Preparing VQC and FT data...")
    all_vqc_dfs = [pd.DataFrame(data).assign(vendor=vendor) for vendor, data in vqc_data.items() if data]
    if all_vqc_dfs:
        df_vqc = pd.concat(all_vqc_dfs, ignore_index=True)
        rename_map = {find_column(df_vqc, ['uid', 'serial']): 'serial_number', find_column(df_vqc, ['status', 'result']): 'vqc_status', find_column(df_vqc, ['reason', 'comments']): 'vqc_reason'}
        df_vqc.rename(columns={k: v for k, v in rename_map.items() if k}, inplace=True)
        if 'serial_number' in df_vqc.columns:
            df_vqc.dropna(subset=['serial_number'], inplace=True)
            df_vqc['serial_number'] = df_vqc['serial_number'].astype(str).str.strip()
            df_vqc = df_vqc[[col for col in ['serial_number', 'vendor', 'vqc_status', 'vqc_reason'] if col in df_vqc.columns]]
    else: df_vqc = pd.DataFrame(columns=['serial_number', 'vendor', 'vqc_status', 'vqc_reason'])
    
    df_ft = pd.DataFrame(ft_data)
    if not df_ft.empty:
        rename_map = {find_column(df_ft, ['uid', 'serial']): 'serial_number', find_column(df_ft, ['status', 'test result']): 'ft_status', find_column(df_ft, ['reason', 'comments']): 'ft_reason'}
        df_ft.rename(columns={k: v for k, v in rename_map.items() if k}, inplace=True)
        if 'serial_number' in df_ft.columns:
            df_ft.dropna(subset=['serial_number'], inplace=True)
            df_ft['serial_number'] = df_ft['serial_number'].astype(str).str.strip()
            df_ft = df_ft[[col for col in ['serial_number', 'ft_status', 'ft_reason'] if col in df_ft.columns]]
    else: df_ft = pd.DataFrame(columns=['serial_number', 'ft_status', 'ft_reason'])

    log_callback("Performing merge...")
    merged_df = pd.merge(df_main, df_vqc, on=['serial_number', 'vendor'], how='left')
    if 'serial_number' in merged_df.columns and 'serial_number' in df_ft.columns:
        merged_df = pd.merge(merged_df, df_ft, on='serial_number', how='left')
    
    merged_df.fillna('', inplace=True)
    
    initial_count = len(merged_df)
    log_callback(f"Successfully merged {initial_count} records. Checking for duplicates...")
    merged_df.drop_duplicates(subset=['serial_number'], keep='last', inplace=True)
    final_count = len(merged_df)
    duplicates_found = initial_count - final_count
    
    if duplicates_found > 0:
        log_callback(f"Removed {duplicates_found} duplicate serial number(s). Final record count: {final_count}.")
    else:
        log_callback("No duplicate serial numbers found.")

    return merged_df.to_dict('records')

@app.route('/api/migrate', methods=['POST'])
def migrate():
    config = request.json
    def generate():
        def log_callback(message):
            yield f"data: {message}\n\n"

        # 1. Connect to Google API
        try:
            yield from log_callback("Connecting to Google API...")
            scopes = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
            creds = Credentials.from_service_account_info(config.get('serviceAccountContent'), scopes=scopes)
            gc = gspread.authorize(creds)
            yield from log_callback("Google API connection successful.")
        except Exception as e:
            yield from log_callback(f"ERROR: Google API connection failed: {e}")
            return

        # 2. Load and Merge Data
        merged_data = []
        try:
            yield from log_callback("Starting parallel data loading from Google Sheets...")
            step7_data, vqc_data, ft_data = [], {}, []
            with ThreadPoolExecutor(max_workers=3) as executor:
                futures = []
                if config.get('vendorDataUrl'): futures.append(executor.submit(_load_sheet_data, 'step7', config, gc, lambda msg: app.logger.info(msg)))
                if config.get('vqcDataUrl'): futures.append(executor.submit(_load_sheet_data, 'vqc', config, gc, lambda msg: app.logger.info(msg)))
                if config.get('ftDataUrl'): futures.append(executor.submit(_load_sheet_data, 'ft', config, gc, lambda msg: app.logger.info(msg)))

                for future in as_completed(futures):
                    try:
                        sheet_type, data = future.result()
                        if sheet_type == 'step7': step7_data = data
                        elif sheet_type == 'vqc': vqc_data = data
                        elif sheet_type == 'ft': ft_data = data
                    except Exception as e:
                        yield from log_callback(f"A task failed during parallel sheet loading: {e}")
            
            yield from log_callback("Parallel data loading complete. Starting merge...")
            merged_data = _merge_ring_data_fast(step7_data, vqc_data, ft_data, lambda msg: app.logger.info(msg))
            yield from log_callback(f"Successfully loaded and merged {len(merged_data)} records.")

        except Exception as e:
            yield from log_callback(f"ERROR: Failed to load or merge data: {e}")
            return

        if not merged_data:
            yield from log_callback("No data to migrate.")
            return

        # 3. Migrate Data
        conn = None
        try:
            conn = get_db_connection()
            with conn.cursor() as cursor:
                yield from log_callback("Creating temporary table for bulk data loading...")
                cursor.execute("""
                    CREATE TEMP TABLE rings_temp (
                        date DATE, mo_number VARCHAR(50), vendor VARCHAR(50), serial_number VARCHAR(100) UNIQUE,
                        ring_size VARCHAR(100), sku VARCHAR(50), vqc_status VARCHAR(100),
                        vqc_reason TEXT, ft_status VARCHAR(100), ft_reason TEXT
                    ) ON COMMIT DROP;
                """)

                yield from log_callback("Preparing data for bulk COPY...")
                string_buffer = io.StringIO()
                cols = ['date', 'mo_number', 'vendor', 'serial_number', 'ring_size', 'sku', 'vqc_status', 'vqc_reason', 'ft_status', 'ft_reason']
                null_identifier = '\\N'

                for record in merged_data:
                    row_data = []
                    for col in cols:
                        value = record.get(col)
                        is_missing = pd.isna(value) or str(value).strip() == ''

                        if col == 'date':
                            if is_missing:
                                clean_value = null_identifier
                            else:
                                try:
                                    clean_value = pd.to_datetime(value).date().isoformat()
                                except (ValueError, TypeError):
                                    clean_value = null_identifier
                        else:
                            if is_missing:
                                clean_value = ''
                            else:
                                clean_value = str(value).replace('\t', ' ').replace('\n', ' ').replace('\r', ' ')
                        
                        row_data.append(clean_value)
                    
                    string_buffer.write('\t'.join(row_data) + '\n')
                
                string_buffer.seek(0)
                
                yield from log_callback(f"Copying {len(merged_data)} records to DB...")
                cursor.copy_expert(f"COPY rings_temp({','.join(cols)}) FROM STDIN WITH (FORMAT text, NULL '{null_identifier}')", string_buffer)

                yield from log_callback("Updating existing records...")
                update_sql = """
                UPDATE rings r SET
                    date = t.date, mo_number = t.mo_number, vendor = t.vendor, ring_size = t.ring_size,
                    sku = t.sku, vqc_status = t.vqc_status, vqc_reason = t.vqc_reason,
                    ft_status = t.ft_status, ft_reason = t.ft_reason, updated_at = CURRENT_TIMESTAMP
                FROM rings_temp t
                WHERE r.serial_number = t.serial_number;
                """
                cursor.execute(update_sql)
                yield from log_callback(f"{cursor.rowcount} existing records updated.")

                yield from log_callback("Inserting new records...")
                insert_sql = """
                INSERT INTO rings (date, mo_number, vendor, serial_number, ring_size, sku, vqc_status, vqc_reason, ft_status, ft_reason)
                SELECT t.date, t.mo_number, t.vendor, t.serial_number, t.ring_size, t.sku, t.vqc_status, t.vqc_reason, t.ft_status, t.ft_reason
                FROM rings_temp t
                LEFT JOIN rings r ON t.serial_number = r.serial_number
                WHERE r.serial_number IS NULL;
                """
                cursor.execute(insert_sql)
                yield from log_callback(f"{cursor.rowcount} new records inserted.")

            conn.commit()
            yield from log_callback("Migration completed successfully!")

        except (psycopg2.Error, Exception) as e:
            if conn: conn.rollback()
            yield from log_callback(f"ERROR: High-speed migration failed: {e}")
        finally:
            if conn: return_db_connection(conn)

    return Response(generate(), mimetype='text/event-stream')



@app.route('/api/search', methods=['POST'])
def search():
    filters = request.json
    app.logger.info(f"Received search filters: {filters}")  # Debug log
    
    query = "SELECT date, vendor, mo_number, serial_number, vqc_status, ft_status, vqc_reason, ft_reason FROM rings WHERE 1=1"
    params = []

    try:
        # Use UPPER for case-insensitive comparison
        if filters.get('serialNumbers'):
            serial_numbers = [s.strip().upper() for s in filters['serialNumbers'].split(',') if s.strip()]
            if serial_numbers:
                query += " AND UPPER(serial_number) = ANY(%s)"
                params.append(serial_numbers)

        if filters.get('moNumbers'):
            mo_numbers = [s.strip().upper() for s in filters['moNumbers'].split(',') if s.strip()]
            if mo_numbers:
                query += " AND UPPER(mo_number) = ANY(%s)"
                params.append(mo_numbers)

        # Improved date filtering with better error handling
        if filters.get('dateFrom'):
            date_from_str = filters['dateFrom']
            app.logger.info(f"Processing dateFrom: {date_from_str}")
            try:
                # Parse the date string - handle both YYYY-MM-DD and other formats
                if isinstance(date_from_str, str):
                    # Try to parse the date
                    from_date = pd.to_datetime(date_from_str).date()
                    query += " AND date >= %s"
                    params.append(from_date)
                    app.logger.info(f"Successfully parsed dateFrom: {from_date}")
                else:
                    app.logger.warning(f"dateFrom is not a string: {type(date_from_str)}")
            except Exception as e:
                app.logger.error(f"Failed to parse dateFrom '{date_from_str}': {e}")
                return jsonify({'error': f'Invalid dateFrom format: {date_from_str}'}), 400

        if filters.get('dateTo'):
            date_to_str = filters['dateTo']
            app.logger.info(f"Processing dateTo: {date_to_str}")
            try:
                # Parse the date string - handle both YYYY-MM-DD and other formats
                if isinstance(date_to_str, str):
                    # Try to parse the date
                    to_date = pd.to_datetime(date_to_str).date()
                    query += " AND date <= %s"
                    params.append(to_date)
                    app.logger.info(f"Successfully parsed dateTo: {to_date}")
                else:
                    app.logger.warning(f"dateTo is not a string: {type(date_to_str)}")
            except Exception as e:
                app.logger.error(f"Failed to parse dateTo '{date_to_str}': {e}")
                return jsonify({'error': f'Invalid dateTo format: {date_to_str}'}), 400
        
        # Handle multi-select filters
        if filters.get('vendor') and len(filters['vendor']) > 0:
            query += " AND vendor = ANY(%s)"
            params.append(filters['vendor'])
            
        if filters.get('vqcStatus') and len(filters['vqcStatus']) > 0:
            query += " AND vqc_status = ANY(%s)"
            params.append(filters['vqcStatus'])
            
        if filters.get('ftStatus') and len(filters['ftStatus']) > 0:
            query += " AND ft_status = ANY(%s)"
            params.append(filters['ftStatus'])
            
        if filters.get('rejectionReason') and len(filters['rejectionReason']) > 0:
            query += " AND (vqc_reason = ANY(%s) OR ft_reason = ANY(%s))"
            params.extend([filters['rejectionReason'], filters['rejectionReason']])

        query += " ORDER BY date DESC, id DESC LIMIT 5000;"
        
        app.logger.info(f"Final query: {query}")
        app.logger.info(f"Query parameters: {params}")

        conn = get_db_connection()
        cur = conn.cursor()
        cur.execute(query, tuple(params))
        
        colnames = [desc[0] for desc in cur.description]
        data = [dict(zip(colnames, row)) for row in cur.fetchall()]
        
        cur.close()
        return_db_connection(conn)
        
        app.logger.info(f"Search completed successfully, returning {len(data)} records")
        return jsonify(data)
        
    except psycopg2.Error as db_err:
        app.logger.error(f"Database error during search: {db_err}")
        return jsonify({'error': f'Database error: {str(db_err)}'}), 500
    except Exception as e:
        app.logger.error(f"Unexpected error during search: {e}")
        return jsonify({'error': f'Search failed: {str(e)}'}), 500
@app.route('/api/search/filters', methods=['GET'])
def get_search_filters():
    """Gets distinct values for search filters from the database."""
    conn = None
    options = {}
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute("SELECT DISTINCT vendor FROM rings WHERE vendor IS NOT NULL AND vendor != '' ORDER BY vendor;")
            options['vendors'] = [row[0] for row in cursor.fetchall()]
            cursor.execute("SELECT DISTINCT vqc_status FROM rings WHERE vqc_status IS NOT NULL AND vqc_status != '' ORDER BY vqc_status;")
            options['vqc_statuses'] = [row[0] for row in cursor.fetchall()]
            cursor.execute("SELECT DISTINCT ft_status FROM rings WHERE ft_status IS NOT NULL AND ft_status != '' ORDER BY ft_status;")
            options['ft_statuses'] = [row[0] for row in cursor.fetchall()]
            reason_query = "SELECT DISTINCT reason FROM (SELECT vqc_reason AS reason FROM rings WHERE vqc_reason IS NOT NULL AND vqc_reason != '' UNION ALL SELECT ft_reason FROM rings WHERE ft_reason IS NOT NULL AND ft_reason != '') AS reasons ORDER BY 1;"
            cursor.execute(reason_query)
            options['reasons'] = [row[0] for row in cursor.fetchall()]
        return jsonify(options)
    except psycopg2.Error as db_err:
        app.logger.error(f"Database error loading filters: {db_err}")
        return jsonify(status="error", message=f"Database error loading filters: {db_err}"), 500
    finally:
        if conn: return_db_connection(conn)

@app.route('/api/search/export', methods=['POST'])
def export_search_results():
    """Exports search results to a CSV file."""
    filters = request.json
    conn = None
    try:
        base_query = "SELECT date, vendor, mo_number, serial_number, vqc_status, ft_status, vqc_reason, ft_reason FROM rings"
        where_clauses, params = [], []

        if filters.get('serialNumbers'):
            serial_numbers = [s.strip().upper() for s in filters['serialNumbers'].split(',') if s.strip()]
            if serial_numbers:
                where_clauses.append("UPPER(serial_number) = ANY(%s)")
                params.append(serial_numbers)
        if filters.get('moNumbers'):
            mo_numbers = [s.strip().upper() for s in filters['moNumbers'].split(',') if s.strip()]
            if mo_numbers:
                where_clauses.append("UPPER(mo_number) = ANY(%s)")
                params.append(mo_numbers)
        if filters.get('dateFrom'):
            where_clauses.append("date >= %s")
            params.append(filters['dateFrom'])
        if filters.get('dateTo'):
            where_clauses.append("date <= %s")
            params.append(filters['dateTo'])
        if filters.get('vendor'):
            where_clauses.append("vendor = ANY(%s)")
            params.append(filters['vendor'])
        if filters.get('vqcStatus'):
            where_clauses.append("vqc_status = ANY(%s)")
            params.append(filters['vqcStatus'])
        if filters.get('ftStatus'):
            where_clauses.append("ft_status = ANY(%s)")
            params.append(filters['ftStatus'])
        if filters.get('rejectionReason'):
            where_clauses.append("(vqc_reason = ANY(%s) OR ft_reason = ANY(%s))")
            params.extend([filters['rejectionReason'], filters['rejectionReason']])

        if where_clauses:
            base_query += " WHERE " + " AND ".join(where_clauses)
        
        base_query += " ORDER BY date DESC, id DESC;"

        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute(base_query, tuple(params))
            results = cursor.fetchall()
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Write headers
            headers = [desc[0] for desc in cursor.description]
            writer.writerow(headers)
            
            # Write data
            for row in results:
                writer.writerow(row)
            
            output.seek(0)
            return Response(output, mimetype="text/csv", headers={"Content-Disposition":"attachment;filename=search_results.csv"})

    except (psycopg2.Error, Exception) as e:
        app.logger.error(f"Export failed: {e}")
        return jsonify(status="error", message=f"Export failed: {e}"), 500
    finally:
        if conn: return_db_connection(conn)


@app.route('/api/test_sheets_connection', methods=['POST'])
def test_sheets_connection():
    config = request.json
    try:
        service_account_info = config.get('serviceAccountContent')
        if not service_account_info or not isinstance(service_account_info, dict):
            return jsonify({'status': 'error', 'message': 'Invalid or missing service account JSON content.'}), 400

        creds = Credentials.from_service_account_info(
            service_account_info,
            scopes=['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        )
        gc = gspread.authorize(creds)
        
        sheet_urls = {
            "Vendor Data": config.get('vendorDataUrl'),
            "VQC Data": config.get('vqcDataUrl'),
            "FT Data": config.get('ftDataUrl')
        }

        # Check if any URLs are provided
        if not any(sheet_urls.values()):
            return jsonify({'status': 'error', 'message': 'No Google Sheet URLs provided for connection test.'}), 400

        overall_status = 'success'
        results = []
        for name, url in sheet_urls.items():
            if not url:
                results.append(f"✗ {name}: No URL provided.")
                overall_status = 'error'
                continue
            try:
                sheet = gc.open_by_url(url)
                results.append(f"✓ {name}: Connected to '{sheet.title}'")
            except Exception as e:
                results.append(f"✗ {name}: FAILED ({e})")
                overall_status = 'error'
        
        return jsonify({'status': overall_status, 'message': "\n".join(results)})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/reports/daily', methods=['POST'])
@app.route('/api/reports/daily', methods=['POST'])
def get_daily_report():
    """Generates a comprehensive daily production report with correct ring status logic."""
    config = request.json
    selected_date = config.get('date')
    selected_vendor = config.get('vendor', 'all')
    
    if not selected_date:
        return jsonify({'error': 'Date is required'}), 400
    
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # Base query conditions
            date_condition = "date = %s"
            vendor_condition = "" if selected_vendor == 'all' else " AND vendor = %s"
            params = [selected_date]
            if selected_vendor != 'all':
                params.append(selected_vendor)
            
            # Get all rings with their VQC and FT data
            cursor.execute(f"""
                SELECT 
                    vendor,
                    serial_number,
                    mo_number,
                    sku,
                    ring_size,
                    vqc_status,
                    vqc_reason,
                    ft_status,
                    ft_reason,
                    created_at
                FROM rings 
                WHERE {date_condition}{vendor_condition}
                ORDER BY created_at, vendor, serial_number
            """, tuple(params))
            
            rings_data = cursor.fetchall()
            
            if not rings_data:
                return jsonify({
                    'date': selected_date,
                    'vendor': selected_vendor,
                    'totalReceived': 0,
                    'totalAccepted': 0,
                    'totalRejected': 0,
                    'totalPending': 0,
                    'yield': 0,
                    'vqcBreakdown': {'accepted': 0, 'rejected': 0, 'pending': 0, 'rejectionReasons': []},
                    'ftBreakdown': {'accepted': 0, 'rejected': 0, 'pending': 0, 'rejectionReasons': []},
                    'hourlyData': [],
                    'vendorBreakdown': []
                })
            
            # Process each ring according to the business logic
            processed_rings = []
            vendor_stats = {}
            vqc_rejection_reasons = {}
            ft_rejection_reasons = {}
            hourly_stats = {}
            
            for ring in rings_data:
                vendor, serial, mo, sku, size, vqc_status, vqc_reason, ft_status, ft_reason, created_at = ring
                
                # Initialize vendor stats if not exists
                if vendor not in vendor_stats:
                    vendor_stats[vendor] = {'received': 0, 'accepted': 0, 'rejected': 0, 'pending': 0}
                
                # Initialize hourly stats
                hour = created_at.hour if created_at else 0
                if hour not in hourly_stats:
                    hourly_stats[hour] = {'received': 0, 'accepted': 0, 'rejected': 0, 'pending': 0}
                
                vendor_stats[vendor]['received'] += 1
                hourly_stats[hour]['received'] += 1
                
                # Apply business logic to determine final status
                has_vqc_data = vqc_status is not None and str(vqc_status).strip() != ''
                has_ft_data = ft_status is not None and str(ft_status).strip() != ''
                
                final_status = 'Pending'
                final_reason = ''
                stage = 'Unknown'
                
                if not has_vqc_data and not has_ft_data:
                    # Neither VQC nor FT data - Pending
                    final_status = 'Pending'
                    stage = 'VQC'
                    
                elif not has_vqc_data and has_ft_data:
                    # No VQC data, but FT data exists - Use FT as final
                    final_status = 'Accepted' if str(ft_status).upper() in ['ACCEPTED', 'PASS'] else 'Rejected'
                    final_reason = ft_reason if final_status == 'Rejected' else ''
                    stage = 'FT'
                    
                elif has_vqc_data and not has_ft_data:
                    # VQC data exists, no FT data - VQC is final (FT pending)
                    final_status = 'Accepted' if str(vqc_status).upper() in ['ACCEPTED', 'PASS'] else 'Rejected'
                    final_reason = vqc_reason if final_status == 'Rejected' else ''
                    stage = 'VQC'
                    
                else:
                    # Both VQC and FT data exist - FT is final (only VQC accepted rings go to FT)
                    final_status = 'Accepted' if str(ft_status).upper() in ['ACCEPTED', 'PASS'] else 'Rejected'
                    final_reason = ft_reason if final_status == 'Rejected' else ''
                    stage = 'FT'
                
                # Update statistics
                if final_status == 'Accepted':
                    vendor_stats[vendor]['accepted'] += 1
                    hourly_stats[hour]['accepted'] += 1
                elif final_status == 'Rejected':
                    vendor_stats[vendor]['rejected'] += 1
                    hourly_stats[hour]['rejected'] += 1
                    
                    # Track rejection reasons
                    if final_reason and final_reason.strip():
                        if stage == 'VQC':
                            vqc_rejection_reasons[final_reason] = vqc_rejection_reasons.get(final_reason, 0) + 1
                        else:
                            ft_rejection_reasons[final_reason] = ft_rejection_reasons.get(final_reason, 0) + 1
                else:
                    vendor_stats[vendor]['pending'] += 1
                    hourly_stats[hour]['pending'] += 1
                
                processed_rings.append({
                    'vendor': vendor,
                    'serial_number': serial,
                    'mo_number': mo,
                    'sku': sku,
                    'ring_size': size,
                    'vqc_status': vqc_status,
                    'vqc_reason': vqc_reason,
                    'ft_status': ft_status,
                    'ft_reason': ft_reason,
                    'final_status': final_status,
                    'final_reason': final_reason,
                    'stage': stage,
                    'created_at': created_at
                })
            
            # Calculate totals
            total_received = len(processed_rings)
            total_accepted = sum(stats['accepted'] for stats in vendor_stats.values())
            total_rejected = sum(stats['rejected'] for stats in vendor_stats.values())
            total_pending = sum(stats['pending'] for stats in vendor_stats.values())
            
            # Calculate yield (excluding pending rings from yield calculation)
            completed_rings = total_accepted + total_rejected
            overall_yield = (total_accepted / completed_rings * 100) if completed_rings > 0 else 0
            
            # Prepare vendor breakdown if 'all' is selected
            vendor_breakdown_data = []
            if selected_vendor == 'all':
                for vendor_name, stats in vendor_stats.items():
                    completed = stats['accepted'] + stats['rejected']
                    vendor_yield = (stats['accepted'] / completed * 100) if completed > 0 else 0
                    vendor_breakdown_data.append({
                        'vendor': vendor_name,
                        'totalReceived': stats['received'],
                        'totalAccepted': stats['accepted'],
                        'totalRejected': stats['rejected'],
                        'totalPending': stats['pending'],
                        'yield': round(vendor_yield, 2)
                    })
            
            # Prepare VQC rejection reasons
            total_vqc_rejected = sum(vqc_rejection_reasons.values())
            vqc_rejection_list = []
            for reason, count in sorted(vqc_rejection_reasons.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_vqc_rejected * 100) if total_vqc_rejected > 0 else 0
                vqc_rejection_list.append({
                    'reason': reason,
                    'count': count,
                    'percentage': round(percentage, 1)
                })
            
            # Prepare FT rejection reasons
            total_ft_rejected = sum(ft_rejection_reasons.values())
            ft_rejection_list = []
            for reason, count in sorted(ft_rejection_reasons.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_ft_rejected * 100) if total_ft_rejected > 0 else 0
                ft_rejection_list.append({
                    'reason': reason,
                    'count': count,
                    'percentage': round(percentage, 1)
                })
            
            # Prepare hourly data
            hourly_data = []
            for hour in sorted(hourly_stats.keys()):
                hourly_data.append({
                    'hour': f"{hour:02d}:00",
                    'received': hourly_stats[hour]['received'],
                    'accepted': hourly_stats[hour]['accepted'],
                    'rejected': hourly_stats[hour]['rejected'],
                    'pending': hourly_stats[hour]['pending']
                })
            
            # Count VQC and FT specific stats for breakdown
            vqc_accepted = len([r for r in processed_rings if r['vqc_status'] and str(r['vqc_status']).upper() in ['ACCEPTED', 'PASS']])
            vqc_rejected = len([r for r in processed_rings if r['vqc_status'] and str(r['vqc_status']).upper() not in ['ACCEPTED', 'PASS']])
            vqc_pending = len([r for r in processed_rings if not r['vqc_status'] or str(r['vqc_status']).strip() == ''])
            
            ft_accepted = len([r for r in processed_rings if r['ft_status'] and str(r['ft_status']).upper() in ['ACCEPTED', 'PASS']])
            ft_rejected = len([r for r in processed_rings if r['ft_status'] and str(r['ft_status']).upper() not in ['ACCEPTED', 'PASS'] and r['ft_status'] and str(r['ft_status']).strip() != ''])
            ft_pending = len([r for r in processed_rings if not r['ft_status'] or str(r['ft_status']).strip() == ''])
            
            return jsonify({
                'date': selected_date,
                'vendor': selected_vendor,
                'totalReceived': total_received,
                'totalAccepted': total_accepted,
                'totalRejected': total_rejected,
                'totalPending': total_pending,
                'yield': round(overall_yield, 2),
                'vqcBreakdown': {
                    'accepted': vqc_accepted,
                    'rejected': vqc_rejected,
                    'pending': vqc_pending,
                    'rejectionReasons': vqc_rejection_list
                },
                'ftBreakdown': {
                    'accepted': ft_accepted,
                    'rejected': ft_rejected,
                    'pending': ft_pending,
                    'rejectionReasons': ft_rejection_list
                },
                'hourlyData': hourly_data,
                'vendorBreakdown': vendor_breakdown_data
            })
        
    except (psycopg2.Error, Exception) as e:
        app.logger.error(f"Error generating daily report: {e}")
        return jsonify({'error': f'Failed to generate report: {str(e)}'}), 500
    finally:
        if conn:
            return_db_connection(conn)

@app.route('/api/reports/export', methods=['POST'])
def export_daily_report():
    """Exports daily report data as CSV or Excel."""
    config = request.json
    selected_date = config.get('date')
    selected_vendor = config.get('vendor', 'all')
    export_format = config.get('format', 'csv')  # 'csv' or 'excel'
    
    if not selected_date:
        return jsonify({'error': 'Date is required'}), 400
    
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # Get detailed data for export
            date_condition = "date = %s"
            vendor_condition = "" if selected_vendor == 'all' else " AND vendor = %s"
            params = [selected_date]
            if selected_vendor != 'all':
                params.append(selected_vendor)
            
            cursor.execute(f"""
                SELECT 
                    date,
                    vendor,
                    serial_number,
                    mo_number,
                    sku,
                    ring_size,
                    vqc_status,
                    vqc_reason,
                    ft_status,
                    ft_reason,
                    CASE 
                        WHEN vqc_status = 'ACCEPTED' THEN 'Accepted'
                        ELSE 'Rejected'
                    END as overall_status,
                    created_at
                FROM rings 
                WHERE {date_condition}{vendor_condition}
                ORDER BY created_at, vendor, serial_number
            """, tuple(params))
            
            results = cursor.fetchall()
            
            if export_format.lower() == 'csv':
                output = io.StringIO()
                writer = csv.writer(output)
                
                # Write headers
                headers = [
                    'Date', 'Vendor', 'Serial Number', 'MO Number', 'SKU', 'Ring Size',
                    'VQC Status', 'VQC Reason', 'FT Status', 'FT Reason', 
                    'Overall Status', 'Created At'
                ]
                writer.writerow(headers)
                
                # Write data
                for row in results:
                    writer.writerow(row)
                
                output.seek(0)
                filename = f"daily_report_{selected_date}_{selected_vendor}.csv"
                
                return Response(
                    output.getvalue(),
                    mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment;filename={filename}"}
                )
            
            elif export_format.lower() == 'excel':
                # For Excel export, you'd typically use pandas or openpyxl
                # This is a simplified version using pandas
                import pandas as pd
                
                df = pd.DataFrame(results, columns=[
                    'Date', 'Vendor', 'Serial Number', 'MO Number', 'SKU', 'Ring Size',
                    'VQC Status', 'VQC Reason', 'FT Status', 'FT Reason', 
                    'Overall Status', 'Created At'
                ])
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Daily Report', index=False)
                
                output.seek(0)
                filename = f"daily_report_{selected_date}_{selected_vendor}.xlsx"
                
                return Response(
                    output.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename={filename}"}
                )
            
            else:
                return jsonify({'error': 'Invalid export format'}), 400
                
    except (psycopg2.Error, Exception) as e:
        app.logger.error(f"Error exporting daily report: {e}")
        return jsonify({'error': f'Failed to export report: {str(e)}'}), 500
    finally:
        if conn:
            return_db_connection(conn)

# Add this endpoint to your app.py file

@app.route('/api/reports/rejection-trends', methods=['POST'])
def get_rejection_trends():
    """Generates rejection trends data in spreadsheet format."""
    config = request.json
    date_from = config.get('dateFrom')
    date_to = config.get('dateTo')
    selected_vendor = config.get('vendor')
    
    if not all([date_from, date_to, selected_vendor]):
        return jsonify({'error': 'dateFrom, dateTo, and vendor are required'}), 400
    
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # Generate date range
            cursor.execute("""
                SELECT generate_series(%s::date, %s::date, '1 day'::interval)::date as date_col
            """, (date_from, date_to))
            date_range = [row[0].strftime('%Y-%m-%d') for row in cursor.fetchall()]
            
            # Get rejection data for the date range and vendor
            cursor.execute("""
                SELECT 
                    date,
                    vendor,
                    vqc_reason,
                    ft_reason,
                    vqc_status,
                    ft_status
                FROM rings 
                WHERE date BETWEEN %s AND %s 
                AND vendor = %s
                AND (
                    (vqc_status IS NOT NULL AND UPPER(vqc_status) NOT IN ('ACCEPTED', 'PASS', '')) 
                    OR 
                    (ft_status IS NOT NULL AND UPPER(ft_status) NOT IN ('ACCEPTED', 'PASS', ''))
                )
            """, (date_from, date_to, selected_vendor))
            
            rejection_records = cursor.fetchall()
            
            # Define rejection categories
            rejection_categories = {
                'ASSEMBLY': [
                    'BLACK GLUE',
                    'ULTRAHUMAN TEXT SMUDGED',
                    'WHITE PATCH ON BATTERY',
                    'WHITE PATCH ON INSERT',
                    'WHITE PATCH ON PCB',
                    'WHITE PATCH ON TAPE NEAR BATTERY',
                    'WRONG RX COIL'
                ],
                'CASTING': [
                    'MICRO BUBBLES',
                    'ALIGNMENT ISSUE',
                    'DENT ON RESIN',
                    'DUST INSIDE RESIN',
                    'RESIN CURING ISSUE',
                    'SHORT FILL OF RESIN',
                    'SPM REJECTION',
                    'TIGHT FIT FOR CHARGE',
                    'LOOSE FITTING ON CHARGER',
                    'RESIN SHRINKAGE',
                    'WRONG MOULD',
                    'GLOP TOP ISSUE'
                ],
                'FUNCTIONAL': [
                    '100% ISSUE',
                    '3 SENSOR ISSUE',
                    'BATTERY ISSUE',
                    'BLUETOOTH HEIGHT ISSUE',
                    'CE TAPE ISSUE',
                    'CHARGING CODE ISSUE',
                    'COIL THICKNESS ISSUE/BATTERY THICKNESS',
                    'COMPONENT HEIGHT ISSUE',
                    'CURRENT ISSUE',
                    'DISCONNECTING ISSUE',
                    'HRS BUBBLE',
                    'HRS COATING HEIGHT ISSUE',
                    'HRS DOUBLE LIGHT ISSUE',
                    'HRS HEIGHT ISSUE',
                    'NO NOTIFICATION IN CDT',
                    'NOT ADVERTISING (WINGLESS PCB)',
                    'NOT CHARGING',
                    'SENSOR ISSUE',
                    'STC ISSUE',
                    'R&D REJECTION'
                ],
                'POLISHING': [
                    'IMPROPER RESIN FINISH',
                    'RESIN DAMAGE',
                    'RX COIL SCRATCH',
                    'SCRATCHES ON RESIN',
                    'SIDE SCRATCH',
                    'SIDE SCRATCH (EMERY)',
                    'SHELL COATING REMOVED',
                    'UNEVEN POLISHING',
                    'WHITE PATCH ON SHELL AFTER POLISHING',
                    'SCRATCHES ON SHELL & SIDE SHELL'
                ],
                'SHELL': [
                    'BLACK MARKS ON SHELL',
                    'DENT ON SHELL',
                    'DISCOLORATION',
                    'IRREGULAR SHELL SHAPE',
                    'SHELL COATING ISSUE',
                    'WHITE MARKS ON SHELL'
                ]
            }
            
            # Create a mapping from rejection reason to stage
            reason_to_stage = {}
            for stage, reasons in rejection_categories.items():
                for reason in reasons:
                    reason_to_stage[reason.upper()] = stage
            
            # Initialize data structure for the spreadsheet format
            trends_data = []
            
            # Process each rejection category
            for stage, rejection_types in rejection_categories.items():
                for rejection_type in rejection_types:
                    row_data = {
                        'stage': stage,
                        'rejection': rejection_type,
                        'dateWiseData': {},
                        'totals': {'total': 0}
                    }
                    
                    # Initialize all dates with 0
                    for date in date_range:
                        row_data['dateWiseData'][date] = 0
                    
                    # Count rejections for this type across all dates
                    for record in rejection_records:
                        record_date, vendor, vqc_reason, ft_reason, vqc_status, ft_status = record
                        record_date_str = record_date.strftime('%Y-%m-%d')
                        
                        # Check if this record matches our rejection type
                        reasons_to_check = []
                        if vqc_reason and vqc_reason.strip():
                            reasons_to_check.append(vqc_reason.strip().upper())
                        if ft_reason and ft_reason.strip():
                            reasons_to_check.append(ft_reason.strip().upper())
                        
                        for reason in reasons_to_check:
                            if reason == rejection_type.upper():
                                if record_date_str in row_data['dateWiseData']:
                                    row_data['dateWiseData'][record_date_str] += 1
                                    row_data['totals']['total'] += 1
                    
                    trends_data.append(row_data)
            
            # Calculate summary statistics
            total_rejections = sum(row['totals']['total'] for row in trends_data)
            stage_wise_totals = {}
            for row in trends_data:
                stage = row['stage']
                if stage not in stage_wise_totals:
                    stage_wise_totals[stage] = 0
                stage_wise_totals[stage] += row['totals']['total']
            
            return jsonify({
                'dateRange': date_range,
                'vendor': selected_vendor,
                'dateFrom': date_from,
                'dateTo': date_to,
                'rejectionData': trends_data,
                'summary': {
                    'totalRejections': total_rejections,
                    'stageWiseTotals': stage_wise_totals,
                    'dateRange': len(date_range)
                }
            })
            
    except (psycopg2.Error, Exception) as e:
        app.logger.error(f"Error generating rejection trends: {e}")
        return jsonify({'error': f'Failed to generate rejection trends: {str(e)}'}), 500
    finally:
        if conn:
            return_db_connection(conn)

@app.route('/api/reports/rejection-trends/export', methods=['POST'])
def export_rejection_trends():
    """Exports rejection trends data as CSV or Excel."""
    config = request.json
    date_from = config.get('dateFrom')
    date_to = config.get('dateTo')
    selected_vendor = config.get('vendor')
    export_format = config.get('format', 'csv')
    
    if not all([date_from, date_to, selected_vendor]):
        return jsonify({'error': 'dateFrom, dateTo, and vendor are required'}), 400
    
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # Generate date range
            cursor.execute("""
                SELECT generate_series(%s::date, %s::date, '1 day'::interval)::date as date_col
            """, (date_from, date_to))
            date_range = [row[0].strftime('%Y-%m-%d') for row in cursor.fetchall()]
            
            # Get rejection data (same logic as above but for export)
            cursor.execute("""
                SELECT 
                    date,
                    vendor,
                    vqc_reason,
                    ft_reason,
                    vqc_status,
                    ft_status
                FROM rings 
                WHERE date BETWEEN %s AND %s 
                AND vendor = %s
                AND (
                    (vqc_status IS NOT NULL AND UPPER(vqc_status) NOT IN ('ACCEPTED', 'PASS', '')) 
                    OR 
                    (ft_status IS NOT NULL AND UPPER(ft_status) NOT IN ('ACCEPTED', 'PASS', ''))
                )
            """, (date_from, date_to, selected_vendor))
            
            rejection_records = cursor.fetchall()
            
            # Same rejection categories and processing logic as above
            rejection_categories = {
                'ASSEMBLY': ['BLACK GLUE', 'ULTRAHUMAN TEXT SMUDGED', 'WHITE PATCH ON BATTERY', 'WHITE PATCH ON INSERT', 'WHITE PATCH ON PCB', 'WHITE PATCH ON TAPE NEAR BATTERY', 'WRONG RX COIL'],
                'CASTING': ['MICRO BUBBLES', 'ALIGNMENT ISSUE', 'DENT ON RESIN', 'DUST INSIDE RESIN', 'RESIN CURING ISSUE', 'SHORT FILL OF RESIN', 'SPM REJECTION', 'TIGHT FIT FOR CHARGE', 'LOOSE FITTING ON CHARGER', 'RESIN SHRINKAGE', 'WRONG MOULD', 'GLOP TOP ISSUE'],
                'FUNCTIONAL': ['100% ISSUE', '3 SENSOR ISSUE', 'BATTERY ISSUE', 'BLUETOOTH HEIGHT ISSUE', 'CE TAPE ISSUE', 'CHARGING CODE ISSUE', 'COIL THICKNESS ISSUE/BATTERY THICKNESS', 'COMPONENT HEIGHT ISSUE', 'CURRENT ISSUE', 'DISCONNECTING ISSUE', 'HRS BUBBLE', 'HRS COATING HEIGHT ISSUE', 'HRS DOUBLE LIGHT ISSUE', 'HRS HEIGHT ISSUE', 'NO NOTIFICATION IN CDT', 'NOT ADVERTISING (WINGLESS PCB)', 'NOT CHARGING', 'SENSOR ISSUE', 'STC ISSUE', 'R&D REJECTION'],
                'POLISHING': ['IMPROPER RESIN FINISH', 'RESIN DAMAGE', 'RX COIL SCRATCH', 'SCRATCHES ON RESIN', 'SIDE SCRATCH', 'SIDE SCRATCH (EMERY)', 'SHELL COATING REMOVED', 'UNEVEN POLISHING', 'WHITE PATCH ON SHELL AFTER POLISHING', 'SCRATCHES ON SHELL & SIDE SHELL'],
                'SHELL': ['BLACK MARKS ON SHELL', 'DENT ON SHELL', 'DISCOLORATION', 'IRREGULAR SHELL SHAPE', 'SHELL COATING ISSUE', 'WHITE MARKS ON SHELL']
            }
            
            # Process data into spreadsheet format
            export_data = []
            headers = ['Stage', 'Rejection Type'] + [pd.to_datetime(date).strftime('%d-%b-%Y') for date in date_range] + ['Total']
            
            for stage, rejection_types in rejection_categories.items():
                for rejection_type in rejection_types:
                    row = [stage, rejection_type]
                    total_count = 0
                    
                    for date in date_range:
                        count = 0
                        for record in rejection_records:
                            record_date, vendor, vqc_reason, ft_reason, vqc_status, ft_status = record
                            record_date_str = record_date.strftime('%Y-%m-%d')
                            
                            if record_date_str == date:
                                reasons_to_check = []
                                if vqc_reason and vqc_reason.strip():
                                    reasons_to_check.append(vqc_reason.strip().upper())
                                if ft_reason and ft_reason.strip():
                                    reasons_to_check.append(ft_reason.strip().upper())
                                
                                if rejection_type.upper() in reasons_to_check:
                                    count += 1
                        
                        row.append(count)
                        total_count += count
                    
                    row.append(total_count)
                    export_data.append(row)
            
            if export_format.lower() == 'csv':
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(headers)
                writer.writerows(export_data)
                output.seek(0)
                
                filename = f"rejection_trends_{date_from}_to_{date_to}_{selected_vendor}.csv"
                return Response(
                    output.getvalue(),
                    mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment;filename={filename}"}
                )
            
            elif export_format.lower() == 'excel':
                df = pd.DataFrame(export_data, columns=headers)
                output = io.BytesIO()
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Rejection Trends', index=False)
                    
                    # Get the workbook and worksheet for formatting
                    workbook = writer.book
                    worksheet = writer.sheets['Rejection Trends']
                    
                    # Apply some basic formatting
                    from openpyxl.styles import PatternFill, Font, Alignment
                    
                    # Header formatting
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    
                    for col in range(1, len(headers) + 1):
                        cell = worksheet.cell(row=1, column=col)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                    
                    # Stage column formatting
                    stage_colors = {
                        'ASSEMBLY': PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
                        'CASTING': PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
                        'FUNCTIONAL': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
                        'POLISHING': PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"),
                        'SHELL': PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                    }
                    
                    for row_idx, row_data in enumerate(export_data, start=2):
                        stage = row_data[0]
                        if stage in stage_colors:
                            for col in range(1, 3):  # Stage and Rejection Type columns
                                worksheet.cell(row=row_idx, column=col).fill = stage_colors[stage]
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30)
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                
                output.seek(0)
                filename = f"rejection_trends_{date_from}_to_{date_to}_{selected_vendor}.xlsx"
                
                return Response(
                    output.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename={filename}"}
                )
            
            else:
                return jsonify({'error': 'Invalid export format'}), 400
                
    except (psycopg2.Error, Exception) as e:
        app.logger.error(f"Error generating rejection trends export: {e}")
        return jsonify({'error': f'Failed to generate rejection trends export: {str(e)}'}), 500
    finally:
        if conn:
            return_db_connection(conn)

if __name__ == '__main__':
    init_db_pool()
    app.run(debug=True)