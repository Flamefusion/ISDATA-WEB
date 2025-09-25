from flask import Blueprint, request, jsonify, Response, current_app
import csv
import io
import psycopg2
import pandas as pd
from app.database import get_db_connection, return_db_connection

report_bp = Blueprint('reports', __name__)

@report_bp.route('/reports/vendors', methods=['GET'])
def get_vendors():
    """Returns a list of unique vendors from the rings table."""
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            cursor.execute("SELECT DISTINCT vendor FROM rings ORDER BY vendor")
            vendors = [row[0] for row in cursor.fetchall()]
            return jsonify(['all'] + vendors)
    except ConnectionError:
        # This is a special case where the database is not connected yet.
        # Return an empty list of vendors so the frontend can handle it.
        return jsonify(['all'])
    except (psycopg2.Error, Exception) as e:
        current_app.logger.error(f"Error fetching vendors: {e}")
        return jsonify({'error': f'Failed to fetch vendors: {str(e)}'}), 500
    finally:
        if conn:
            return_db_connection(conn)

@report_bp.route('/reports/daily', methods=['POST'])
def get_daily_report():
    """Generates a comprehensive daily production report with correct ring status logic."""
    config = request.json
    selected_date = config.get('date')
    selected_vendor = config.get('vendor', 'all')
    
    if not selected_date:
        return jsonify({'error': 'Date is required'}), 400
    
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # Base query conditions
            date_condition = "date = %s"
            vendor_condition = "" if selected_vendor == 'all' else " AND vendor = %s"
            params = [selected_date]
            if selected_vendor != 'all':
                params.append(selected_vendor)
            
            # Get all rings with their VQC and FT data
            cursor.execute(f"""
                SELECT 
                    vendor,
                    serial_number,
                    mo_number,
                    sku,
                    ring_size,
                    vqc_status,
                    vqc_reason,
                    ft_status,
                    ft_reason,
                    created_at
                FROM rings 
                WHERE {date_condition}{vendor_condition}
                ORDER BY created_at, vendor, serial_number
            """, tuple(params))
            
            rings_data = cursor.fetchall()
            
            if not rings_data:
                return jsonify({
                    'date': selected_date,
                    'vendor': selected_vendor,
                    'totalReceived': 0,
                    'totalAccepted': 0,
                    'totalRejected': 0,
                    'totalPending': 0,
                    'yield': 0,
                    'vqcBreakdown': {'accepted': 0, 'rejected': 0, 'pending': 0, 'rejectionReasons': []},
                    'ftBreakdown': {'accepted': 0, 'rejected': 0, 'pending': 0, 'rejectionReasons': []},
                    'hourlyData': [],
                    'vendorBreakdown': []
                })
            
            # Process each ring according to the business logic
            processed_rings = []
            vendor_stats = {}
            vqc_rejection_reasons = {}
            ft_rejection_reasons = {}
            hourly_stats = {}
            
            for ring in rings_data:
                vendor, serial, mo, sku, size, vqc_status, vqc_reason, ft_status, ft_reason, created_at = ring
                
                # Initialize vendor stats if not exists
                if vendor not in vendor_stats:
                    vendor_stats[vendor] = {'received': 0, 'accepted': 0, 'rejected': 0, 'pending': 0}
                
                # Initialize hourly stats
                hour = created_at.hour if created_at else 0
                if hour not in hourly_stats:
                    hourly_stats[hour] = {'received': 0, 'accepted': 0, 'rejected': 0, 'pending': 0}
                
                vendor_stats[vendor]['received'] += 1
                hourly_stats[hour]['received'] += 1
                
                # Apply business logic to determine final status
                has_vqc_data = vqc_status is not None and str(vqc_status).strip() != ''
                has_ft_data = ft_status is not None and str(ft_status).strip() != ''
                
                final_status = 'Pending'
                final_reason = ''
                stage = 'Unknown'
                
                if not has_vqc_data and not has_ft_data:
                    # Neither VQC nor FT data - Pending
                    final_status = 'Pending'
                    stage = 'VQC'
                    
                elif not has_vqc_data and has_ft_data:
                    # No VQC data, but FT data exists - Use FT as final
                    final_status = 'Accepted' if str(ft_status).upper() in ['ACCEPTED', 'PASS'] else 'Rejected'
                    final_reason = ft_reason if final_status == 'Rejected' else ''
                    stage = 'FT'
                    
                elif has_vqc_data and not has_ft_data:
                    # VQC data exists, no FT data - VQC is final (FT pending)
                    final_status = 'Accepted' if str(vqc_status).upper() in ['ACCEPTED', 'PASS'] else 'Rejected'
                    final_reason = vqc_reason if final_status == 'Rejected' else ''
                    stage = 'VQC'
                    
                else:
                    # Both VQC and FT data exist - FT is final (only VQC accepted rings go to FT)
                    final_status = 'Accepted' if str(ft_status).upper() in ['ACCEPTED', 'PASS'] else 'Rejected'
                    final_reason = ft_reason if final_status == 'Rejected' else ''
                    stage = 'FT'
                
                # Update statistics
                if final_status == 'Accepted':
                    vendor_stats[vendor]['accepted'] += 1
                    hourly_stats[hour]['accepted'] += 1
                elif final_status == 'Rejected':
                    vendor_stats[vendor]['rejected'] += 1
                    hourly_stats[hour]['rejected'] += 1
                    
                    # Track rejection reasons
                    if final_reason and final_reason.strip():
                        if stage == 'VQC':
                            vqc_rejection_reasons[final_reason] = vqc_rejection_reasons.get(final_reason, 0) + 1
                        else:
                            ft_rejection_reasons[final_reason] = ft_rejection_reasons.get(final_reason, 0) + 1
                else:
                    vendor_stats[vendor]['pending'] += 1
                    hourly_stats[hour]['pending'] += 1
                
                processed_rings.append({
                    'vendor': vendor,
                    'serial_number': serial,
                    'mo_number': mo,
                    'sku': sku,
                    'ring_size': size,
                    'vqc_status': vqc_status,
                    'vqc_reason': vqc_reason,
                    'ft_status': ft_status,
                    'ft_reason': ft_reason,
                    'final_status': final_status,
                    'final_reason': final_reason,
                    'stage': stage,
                    'created_at': created_at
                })
            
            # Calculate totals
            total_received = len(processed_rings)
            total_accepted = sum(stats['accepted'] for stats in vendor_stats.values())
            total_rejected = sum(stats['rejected'] for stats in vendor_stats.values())
            total_pending = sum(stats['pending'] for stats in vendor_stats.values())
            
            # Calculate yield (excluding pending rings from yield calculation)
            completed_rings = total_accepted + total_rejected
            overall_yield = (total_accepted / completed_rings * 100) if completed_rings > 0 else 0
            
            # Prepare vendor breakdown if 'all' is selected
            vendor_breakdown_data = []
            if selected_vendor == 'all':
                for vendor_name, stats in vendor_stats.items():
                    completed = stats['accepted'] + stats['rejected']
                    vendor_yield = (stats['accepted'] / completed * 100) if completed > 0 else 0
                    vendor_breakdown_data.append({
                        'vendor': vendor_name,
                        'totalReceived': stats['received'],
                        'totalAccepted': stats['accepted'],
                        'totalRejected': stats['rejected'],
                        'totalPending': stats['pending'],
                        'yield': round(vendor_yield, 2)
                    })
            
            # Prepare VQC rejection reasons
            total_vqc_rejected = sum(vqc_rejection_reasons.values())
            vqc_rejection_list = []
            for reason, count in sorted(vqc_rejection_reasons.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_vqc_rejected * 100) if total_vqc_rejected > 0 else 0
                vqc_rejection_list.append({
                    'reason': reason,
                    'count': count,
                    'percentage': round(percentage, 1)
                })
            
            # Prepare FT rejection reasons
            total_ft_rejected = sum(ft_rejection_reasons.values())
            ft_rejection_list = []
            for reason, count in sorted(ft_rejection_reasons.items(), key=lambda x: x[1], reverse=True):
                percentage = (count / total_ft_rejected * 100) if total_ft_rejected > 0 else 0
                ft_rejection_list.append({
                    'reason': reason,
                    'count': count,
                    'percentage': round(percentage, 1)
                })
            
            # Prepare hourly data
            hourly_data = []
            for hour in sorted(hourly_stats.keys()):
                hourly_data.append({
                    'hour': f"{hour:02d}:00",
                    'received': hourly_stats[hour]['received'],
                    'accepted': hourly_stats[hour]['accepted'],
                    'rejected': hourly_stats[hour]['rejected'],
                    'pending': hourly_stats[hour]['pending']
                })
            
            # Count VQC and FT specific stats for breakdown
            vqc_accepted = len([r for r in processed_rings if r['vqc_status'] and str(r['vqc_status']).upper() in ['ACCEPTED', 'PASS']])
            vqc_rejected = len([r for r in processed_rings if r['vqc_status'] and str(r['vqc_status']).upper() not in ['ACCEPTED', 'PASS']])
            vqc_pending = len([r for r in processed_rings if not r['vqc_status'] or str(r['vqc_status']).strip() == ''])
            
            ft_accepted = len([r for r in processed_rings if r['ft_status'] and str(r['ft_status']).upper() in ['ACCEPTED', 'PASS']])
            ft_rejected = len([r for r in processed_rings if r['ft_status'] and str(r['ft_status']).upper() not in ['ACCEPTED', 'PASS'] and r['ft_status'] and str(r['ft_status']).strip() != ''])
            ft_pending = len([r for r in processed_rings if not r['ft_status'] or str(r['ft_status']).strip() == ''])
            
            return jsonify({
                'date': selected_date,
                'vendor': selected_vendor,
                'totalReceived': total_received,
                'totalAccepted': total_accepted,
                'totalRejected': total_rejected,
                'totalPending': total_pending,
                'yield': round(overall_yield, 2),
                'vqcBreakdown': {
                    'accepted': vqc_accepted,
                    'rejected': vqc_rejected,
                    'pending': vqc_pending,
                    'rejectionReasons': vqc_rejection_list
                },
                'ftBreakdown': {
                    'accepted': ft_accepted,
                    'rejected': ft_rejected,
                    'pending': ft_pending,
                    'rejectionReasons': ft_rejection_list
                },
                'hourlyData': hourly_data,
                'vendorBreakdown': vendor_breakdown_data
            })
        
    except (psycopg2.Error, Exception) as e:
        current_app.logger.error(f"Error generating daily report: {e}")
        return jsonify({'error': f'Failed to generate report: {str(e)}'}), 500
    finally:
        if conn:
            return_db_connection(conn)

@report_bp.route('/reports/export', methods=['POST'])
def export_daily_report():
    """Exports daily report data as CSV or Excel."""
    config = request.json
    selected_date = config.get('date')
    selected_vendor = config.get('vendor', 'all')
    export_format = config.get('format', 'csv')  # 'csv' or 'excel'
    
    if not selected_date:
        return jsonify({'error': 'Date is required'}), 400
    
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # Get detailed data for export
            date_condition = "date = %s"
            vendor_condition = "" if selected_vendor == 'all' else " AND vendor = %s"
            params = [selected_date]
            if selected_vendor != 'all':
                params.append(selected_vendor)
            
            cursor.execute(f"""
                SELECT 
                    date,
                    vendor,
                    serial_number,
                    mo_number,
                    sku,
                    ring_size,
                    vqc_status,
                    vqc_reason,
                    ft_status,
                    ft_reason,
                    CASE 
                        WHEN vqc_status = 'ACCEPTED' THEN 'Accepted'
                        ELSE 'Rejected'
                    END as overall_status,
                    created_at
                FROM rings 
                WHERE {date_condition}{vendor_condition}
                ORDER BY created_at, vendor, serial_number
            """, tuple(params))
            
            results = cursor.fetchall()
            
            if export_format.lower() == 'csv':
                output = io.StringIO()
                writer = csv.writer(output)
                
                # Write headers
                headers = [
                    'Date', 'Vendor', 'Serial Number', 'MO Number', 'SKU', 'Ring Size',
                    'VQC Status', 'VQC Reason', 'FT Status', 'FT Reason', 
                    'Overall Status', 'Created At'
                ]
                writer.writerow(headers)
                
                # Write data
                for row in results:
                    writer.writerow(row)
                
                output.seek(0)
                filename = f"daily_report_{selected_date}_{selected_vendor}.csv"
                
                return Response(
                    output.getvalue(),
                    mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment;filename={filename}"}
                )
            
            elif export_format.lower() == 'excel':
                # For Excel export using pandas
                df = pd.DataFrame(results, columns=[
                    'Date', 'Vendor', 'Serial Number', 'MO Number', 'SKU', 'Ring Size',
                    'VQC Status', 'VQC Reason', 'FT Status', 'FT Reason', 
                    'Overall Status', 'Created At'
                ])
                
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Daily Report', index=False)
                
                output.seek(0)
                filename = f"daily_report_{selected_date}_{selected_vendor}.xlsx"
                
                return Response(
                    output.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename={filename}"}
                )
            
            else:
                return jsonify({'error': 'Invalid export format'}), 400
                
    except (psycopg2.Error, Exception) as e:
        current_app.logger.error(f"Error exporting daily report: {e}")
        return jsonify({'error': f'Failed to export report: {str(e)}'}), 500
    finally:
        if conn:
            return_db_connection(conn)

@report_bp.route('/reports/rejection-trends', methods=['POST'])
def get_rejection_trends():
    """Generates rejection trends data in spreadsheet format with stage filtering."""
    config = request.json
    date_from = config.get('dateFrom')
    date_to = config.get('dateTo')
    selected_vendor = config.get('vendor')
    rejection_stage_filter = config.get('rejectionStage', 'both')  # 'vqc', 'ft', or 'both'

    if not all([date_from, date_to, selected_vendor]):
        return jsonify({'error': 'dateFrom, dateTo, and vendor are required'}), 400

    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # Generate date range
            cursor.execute("""
                SELECT generate_series(%s::date, %s::date, '1 day'::interval)::date as date_col
            """, (date_from, date_to))
            date_range = [row[0].strftime('%Y-%m-%d') for row in cursor.fetchall()]

            # Fetch all potentially rejected rings
            cursor.execute("""
                SELECT date, vqc_status, vqc_reason, ft_status, ft_reason
                FROM rings
                WHERE date BETWEEN %s AND %s AND vendor = %s
                AND (
                    (vqc_status IS NOT NULL AND UPPER(vqc_status) NOT IN ('ACCEPTED', 'PASS', '')) OR
                    (ft_status IS NOT NULL AND UPPER(ft_status) NOT IN ('ACCEPTED', 'PASS', ''))
                )
            """, (date_from, date_to, selected_vendor))

            raw_records = cursor.fetchall()
            
            # Process records based on business logic to get final rejection reason
            processed_rejections = []
            for record in raw_records:
                date, vqc_status, vqc_reason, ft_status, ft_reason = record
                
                is_vqc_rejected = vqc_status and vqc_status.upper() not in ['ACCEPTED', 'PASS', '']
                is_ft_rejected = ft_status and ft_status.upper() not in ['ACCEPTED', 'PASS', '']

                final_reason = None
                final_stage = None

                # VQC rejection is final and takes precedence
                if is_vqc_rejected:
                    final_stage = 'vqc'
                    final_reason = vqc_reason
                # If not VQC rejected, check for FT rejection
                elif is_ft_rejected:
                    final_stage = 'ft'
                    final_reason = ft_reason
                
                # If a final rejection was determined, and it matches the filter, add it
                if final_stage and final_reason and (rejection_stage_filter == 'both' or rejection_stage_filter == final_stage):
                    processed_rejections.append({
                        'date': date.strftime('%Y-%m-%d'),
                        'reason': final_reason.strip().upper()
                    })

            # Define rejection categories
            rejection_categories = {
                'ASSEMBLY': ['BLACK GLUE', 'ULTRAHUMAN TEXT SMUDGED', 'WHITE PATCH ON BATTERY', 'WHITE PATCH ON BLACK TAPE', 'WHITE PATCH ON PCB', 'WHITE PATCH ON TAPE NEAR BATTERY', 'WRONG RX COIL'],
                'CASTING': ['MICRO BUBBLES', 'ALIGNMENT ISSUE', 'DENT ON RESIN', 'DUST INSIDE RESIN', 'RESIN CURING ISSUE', 'SHORT FILL OF RESIN', 'SPM REJECTION', 'TIGHT FIT FOR CHARGE', 'LOOSE FITTING ON CHARGER', 'RESIN SHRINKAGE', 'WRONG MOULD', 'GLOB TOP ISSUE'],
                'FUNCTIONAL': ['100% ISSUE', '3 SENSOR ISSUE', 'BATTERY ISSUE', 'BLUETOOTH HEIGHT ISSUE', 'CE TAPE ISSUE', 'CHARGING CODE ISSUE', 'COIL THICKNESS ISSUE/BATTERY THICKNESS', 'COMPONENT HEIGHT ISSUE', 'CURRENT ISSUE', 'DISCONNECTING ISSUE', 'HRS BUBBLE', 'HRS COATING HEIGHT ISSUE', 'HRS DOUBLE LIGHT ISSUE', 'HRS HEIGHT ISSUE', 'NO NOTIFICATION IN CDT', 'NOT ADVERTISING (WINGLESS PCB)', 'NOT CHARGING', 'SENSOR ISSUE', 'STC ISSUE', 'R & D REJECTION'],
                'POLISHING': ['IMPROPER RESIN FINISH', 'RESIN DAMAGE', 'RX COIL SCRATCH', 'SCRATCHES ON RESIN', 'SIDE SCRATCH', 'SIDE SCRATCH (EMERY)', 'SHELL COATING REMOVED', 'UNEVEN POLISHING', 'WHITE PATCH ON SHELL AFTER POLISHING', 'SCRATCHES ON SHELL & SIDE SHELL'],
                'SHELL': ['BLACK MARKS ON SHELL', 'DENT ON SHELL', 'DISCOLORATION', 'IRREGULAR SHELL SHAPE', 'SHELL COATING ISSUE', 'WHITE MARKS ON SHELL']
            }

            trends_data = []
            for stage, rejection_types in rejection_categories.items():
                for rejection_type in rejection_types:
                    row_data = {
                        'stage': stage,
                        'rejection': rejection_type,
                        'dateWiseData': {date: 0 for date in date_range},
                        'totals': {'total': 0}
                    }
                    
                    for rejection in processed_rejections:
                        if rejection['reason'] == rejection_type.upper():
                            rejection_date = rejection['date']
                            if rejection_date in row_data['dateWiseData']:
                                row_data['dateWiseData'][rejection_date] += 1
                                row_data['totals']['total'] += 1
                    
                    trends_data.append(row_data)

            # Calculate summary statistics
            total_rejections = sum(row['totals']['total'] for row in trends_data)
            stage_wise_totals = {}
            for row in trends_data:
                stage = row['stage']
                stage_wise_totals[stage] = stage_wise_totals.get(stage, 0) + row['totals']['total']

            return jsonify({
                'dateRange': date_range,
                'vendor': selected_vendor,
                'dateFrom': date_from,
                'dateTo': date_to,
                'rejectionData': trends_data,
                'summary': {
                    'totalRejections': total_rejections,
                    'stageWiseTotals': stage_wise_totals,
                    'dateRange': len(date_range)
                }
            })
            
    except (psycopg2.Error, Exception) as e:
        current_app.logger.error(f"Error generating rejection trends: {e}", exc_info=True)
        return jsonify({'error': f'Failed to generate rejection trends: {str(e)}'}), 500
    finally:
        if conn:
            return_db_connection(conn)

@report_bp.route('/reports/rejection-trends/export', methods=['POST'])
def export_rejection_trends():
    """Exports rejection trends data as CSV or Excel."""
    config = request.json
    date_from = config.get('dateFrom')
    date_to = config.get('dateTo')
    selected_vendor = config.get('vendor')
    export_format = config.get('format', 'csv')
    rejection_stage = config.get('rejectionStage', 'both')
    
    if not all([date_from, date_to, selected_vendor]):
        return jsonify({'error': 'dateFrom, dateTo, and vendor are required'}), 400
    
    conn = None
    try:
        conn = get_db_connection()
        with conn.cursor() as cursor:
            # Generate date range
            cursor.execute("""
                SELECT generate_series(%s::date, %s::date, '1 day'::interval)::date as date_col
            """, (date_from, date_to))
            date_range = [row[0].strftime('%Y-%m-%d') for row in cursor.fetchall()]

            # Dynamically build the rejection condition based on the selected stage
            rejection_condition = ""
            if rejection_stage == 'vqc':
                rejection_condition = "AND (vqc_status IS NOT NULL AND UPPER(vqc_status) NOT IN ('ACCEPTED', 'PASS', ''))"
            elif rejection_stage == 'ft':
                rejection_condition = "AND (ft_status IS NOT NULL AND UPPER(ft_status) NOT IN ('ACCEPTED', 'PASS', ''))"
            else: # 'both'
                rejection_condition = """
                    AND (
                        (vqc_status IS NOT NULL AND UPPER(vqc_status) NOT IN ('ACCEPTED', 'PASS', '')) 
                        OR 
                        (ft_status IS NOT NULL AND UPPER(ft_status) NOT IN ('ACCEPTED', 'PASS', ''))
                    )
                """
            
            # Get rejection data using the dynamic condition
            query = f"""
                SELECT 
                    date,
                    vendor,
                    vqc_reason,
                    ft_reason,
                    vqc_status,
                    ft_status
                FROM rings 
                WHERE date BETWEEN %s AND %s 
                AND vendor = %s
                {rejection_condition}
            """
            cursor.execute(query, (date_from, date_to, selected_vendor))
            
            rejection_records = cursor.fetchall()
            
            # Same rejection categories as above
            rejection_categories = {
                'ASSEMBLY': ['BLACK GLUE', 'ULTRAHUMAN TEXT SMUDGED', 'WHITE PATCH ON BATTERY', 'WHITE PATCH ON INSERT', 'WHITE PATCH ON PCB', 'WHITE PATCH ON TAPE NEAR BATTERY', 'WRONG RX COIL'],
                'CASTING': ['MICRO BUBBLES', 'ALIGNMENT ISSUE', 'DENT ON RESIN', 'DUST INSIDE RESIN', 'RESIN CURING ISSUE', 'SHORT FILL OF RESIN', 'SPM REJECTION', 'TIGHT FIT FOR CHARGE', 'LOOSE FITTING ON CHARGER', 'RESIN SHRINKAGE', 'WRONG MOULD', 'GLOP TOP ISSUE'],
                'FUNCTIONAL': ['100% ISSUE', '3 SENSOR ISSUE', 'BATTERY ISSUE', 'BLUETOOTH HEIGHT ISSUE', 'CE TAPE ISSUE', 'CHARGING CODE ISSUE', 'COIL THICKNESS ISSUE/BATTERY THICKNESS', 'COMPONENT HEIGHT ISSUE', 'CURRENT ISSUE', 'DISCONNECTING ISSUE', 'HRS BUBBLE', 'HRS COATING HEIGHT ISSUE', 'HRS DOUBLE LIGHT ISSUE', 'HRS HEIGHT ISSUE', 'NO NOTIFICATION IN CDT', 'NOT ADVERTISING (WINGLESS PCB)', 'NOT CHARGING', 'SENSOR ISSUE', 'STC ISSUE', 'R&D REJECTION'],
                'POLISHING': ['IMPROPER RESIN FINISH', 'RESIN DAMAGE', 'RX COIL SCRATCH', 'SCRATCHES ON RESIN', 'SIDE SCRATCH', 'SIDE SCRATCH (EMERY)', 'SHELL COATING REMOVED', 'UNEVEN POLISHING', 'WHITE PATCH ON SHELL AFTER POLISHING', 'SCRATCHES ON SHELL & SIDE SHELL'],
                'SHELL': ['BLACK MARKS ON SHELL', 'DENT ON SHELL', 'DISCOLORATION', 'IRREGULAR SHELL SHAPE', 'SHELL COATING ISSUE', 'WHITE MARKS ON SHELL']
            }
            
            # Process data into spreadsheet format
            export_data = []
            headers = ['Stage', 'Rejection Type'] + [pd.to_datetime(date).strftime('%d-%b-%Y') for date in date_range] + ['Total']
            
            for stage, rejection_types in rejection_categories.items():
                for rejection_type in rejection_types:
                    row = [stage, rejection_type]
                    total_count = 0
                    
                    for date in date_range:
                        count = 0
                        for record in rejection_records:
                            record_date, vendor, vqc_reason, ft_reason, vqc_status, ft_status = record
                            record_date_str = record_date.strftime('%Y-%m-%d')
                            
                            if record_date_str == date:
                                reasons_to_check = []
                                if rejection_stage in ['vqc', 'both'] and vqc_reason and vqc_reason.strip():
                                    reasons_to_check.append(vqc_reason.strip().upper())
                                if rejection_stage in ['ft', 'both'] and ft_reason and ft_reason.strip():
                                    reasons_to_check.append(ft_reason.strip().upper())
                                
                                if rejection_type.upper() in reasons_to_check:
                                    count += 1
                        
                        row.append(count)
                        total_count += count
                    
                    row.append(total_count)
                    export_data.append(row)
            
            if export_format.lower() == 'csv':
                output = io.StringIO()
                writer = csv.writer(output)
                writer.writerow(headers)
                writer.writerows(export_data)
                output.seek(0)
                
                filename = f"rejection_trends_{date_from}_to_{date_to}_{selected_vendor}.csv"
                return Response(
                    output.getvalue(),
                    mimetype="text/csv",
                    headers={"Content-Disposition": f"attachment;filename={filename}"}
                )
            
            elif export_format.lower() == 'excel':
                df = pd.DataFrame(export_data, columns=headers)
                output = io.BytesIO()
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Rejection Trends', index=False)
                    
                    workbook = writer.book
                    worksheet = writer.sheets['Rejection Trends']
                    
                    try:
                        from openpyxl.styles import PatternFill, Font, Alignment
                        
                        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        header_font = Font(color="FFFFFF", bold=True)
                        
                        for col in range(1, len(headers) + 1):
                            cell = worksheet.cell(row=1, column=col)
                            cell.fill = header_fill
                            cell.font = header_font
                            cell.alignment = Alignment(horizontal="center")
                        
                        stage_colors = {
                            'ASSEMBLY': PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid"),
                            'CASTING': PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid"),
                            'FUNCTIONAL': PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"),
                            'POLISHING': PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"),
                            'SHELL': PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
                        }
                        
                        for row_idx, row_data in enumerate(export_data, start=2):
                            stage = row_data[0]
                            if stage in stage_colors:
                                for col in range(1, 3):
                                    worksheet.cell(row=row_idx, column=col).fill = stage_colors[stage]
                        
                        for column in worksheet.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 30)
                            worksheet.column_dimensions[column_letter].width = adjusted_width
                    except ImportError:
                        # If openpyxl styles are not available, just create basic Excel
                        pass
                
                output.seek(0)
                filename = f"rejection_trends_{date_from}_to_{date_to}_{selected_vendor}.xlsx"
                
                return Response(
                    output.getvalue(),
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": f"attachment;filename={filename}"}
                )
            
            else:
                return jsonify({'error': 'Invalid export format'}), 400
                
    except (psycopg2.Error, Exception) as e:
        current_app.logger.error(f"Error generating rejection trends export: {e}")
        return jsonify({'error': f'Failed to generate rejection trends export: {str(e)}'}), 500
    finally:
        if conn:
            return_db_connection(conn)